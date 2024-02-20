import pandas as pd
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchWindowException
import time
from credenciais import segredos

caminhoBD = r"C:\Users\naelm\Downloads\Telegram Desktop\bd2.xlsx"
primeiraLinha = 2
codigo = segredos.get('codigo')
senha = segredos.get('senha')
urlSessaoBruto = segredos.get('urlSessaoBruto')
comecarComTitular = 1 # "1" para sim / "1" para dependente

# Abrir o Chrome e fazer login
print('\nAutomação iniciada\n')
navegador = webdriver.Chrome()
if not urlSessaoBruto:
    navegador.get('https://sigo.sh.srv.br/pls/webmin/webnewcadastrousuario.login')
    navegador.find_element("xpath",'//*[@id="webnewcadastrousuario"]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td[2]/input').send_keys(codigo)
    navegador.find_element("xpath",'//*[@id="webnewcadastrousuario"]/table/tbody/tr[1]/td[2]/table/tbody/tr[3]/td[2]/input').click()
    time.sleep(2)
    WebDriverWait(navegador, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="pSenha"]')))
    navegador.find_element("xpath", '//*[@id="pSenha"]').send_keys(senha)
    navegador.find_element("xpath", '//*[@id="Prosseguir"]').click()
    WebDriverWait(navegador, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div/div/table[2]/tbody/tr/td/p[5]/input')))
    urlSessaoBruto = str(navegador.current_url)

pPessoa, pNoCache, pIdSessao = urlSessaoBruto.split('?')[1].split('&')

#Coleta o CSV com Lista de Usuários Ativos
caminho_arquivo = 'C:/Users/naelm/Downloads/Telegram Desktop/empresa.csv'
dados = pd.read_csv(caminho_arquivo, index_col=False, sep=';', dtype={'CPF': str})
df_soTitulares = dados[dados['TIPO USUARIO'] == 'TITULAR']
df_soTitulares.loc[:, 'CPF'] = df_soTitulares['CPF'].str.zfill(11)
colunas_desejadas = ["MATRICULA", "USUARIO", "CODIGO USUARIO", "CPF"]
df_resultado = df_soTitulares[colunas_desejadas]

# Coleta as informacoes da planilha
planilha = load_workbook(caminhoBD)
aba_ativa = planilha.active

linhasExecutadas = 0
linhasErro = 0

try:
    while comecarComTitular <= 2:
        for linha in aba_ativa.iter_rows(min_row=primeiraLinha, max_row=aba_ativa.max_row):
            celula = linha[0]
            status = aba_ativa[f"K{celula.row}"].value
            if celula.value is not None and status != 'Ativo':
                nome = aba_ativa[f"D{celula.row}"].value
                dtNascimento = aba_ativa[f"E{celula.row}"].value
                sexo = aba_ativa[f"F{celula.row}"].value
                cpfTitular = aba_ativa[f"G{celula.row}"].value
                cpfReal = aba_ativa[f"H{celula.row}"].value
                dependencia = aba_ativa[f"I{celula.row}"].value
                dataCasamento = aba_ativa[f"J{celula.row}"].value
                nomeMae = aba_ativa[f"L{celula.row}"].value
                cep = str(aba_ativa[f"Y{celula.row}"].value)
                logradouroNumero = aba_ativa[f"V{celula.row}"].value
                estadoCivil = aba_ativa[f"AB{celula.row}"].value

                # Verifica se CPF está na Lista de Usuários Ativos
                filtro_cpf = df_resultado['CPF'] == cpfTitular
                if filtro_cpf.any():
                    codigo_usuarioId = df_resultado.loc[filtro_cpf, 'CODIGO USUARIO'].values[0]
                    codigo_usuarioNome = df_resultado.loc[filtro_cpf, 'USUARIO'].values[0]
                else:
                    codigo_usuarioId = ''
                    codigo_usuarioNome = ''

                vida = f'Vida linha: {celula.row}, CPF: {cpfReal}, Titular {cpfTitular}'
                if comecarComTitular == 1 :
                    if dependencia == 1:
                        if len(codigo_usuarioId) == 0:
                            print(vida + ' || Iniciando titular')
                            navegador.get('https://sigo.sh.srv.br/pls/webmin/webNewCadastroUsuario.inclusaoContratoFicha?' +
                                          pPessoa + '&' +
                                          pIdSessao + '&' +
                                          'pCpfTitular=' + cpfReal)
                            WebDriverWait(navegador, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="id"]')))
                            navegador.find_element("xpath", '//*[@id="id"]').click()
                            navegador.find_element("xpath", '// *[ @ id = "pNomeTitular"]').send_keys(nome)
                            time.sleep(1)
                            navegador.find_element("xpath", '// *[ @ id = "pCep"]').click()
                            navegador.find_element("xpath", '// *[ @ id = "pCep"]').send_keys(cep)
                            navegador.find_element("xpath", '// *[ @ id = "pCep"]').send_keys(Keys.TAB)
                            try:
                                WebDriverWait(navegador, 5).until(
                                    lambda driver: navegador.find_element("xpath", '// *[ @ id = "pUf"]').get_attribute(
                                        'value') != '')
                                time.sleep(2)
                            except TimeoutException:
                                print('Status: CEP não encontrado na base')
                                linhasErro += 1
                            navegador.find_element("xpath", '// *[ @ id = "pNumero"]').send_keys(logradouroNumero)

                            # Grava a confirmacao na planilha
                            aba_ativa[f"K{celula.row}"] = "Ativo"
                            planilha.save(caminhoBD)
                            print('Status: Finalizada \n')
                            linhasExecutadas += 1
                            time.sleep(5)
                        else:
                            # Grava a confirmacao na planilha
                            aba_ativa[f"K{celula.row}"] = "Ativo"
                            planilha.save(caminhoBD)
                            linhasExecutadas += 1
                            time.sleep(5)
                            print(vida + ' || Titular ativo\n')
                elif comecarComTitular == 2:
                    if len(codigo_usuarioId) > 0:
                        if dependencia != 1:
                            print(vida + ' || Iniciando dependente')
                            navegador.get('https://sigo.sh.srv.br/pls/webmin/webNewCadastroUsuario.InclusaoDependenteFicha?' +
                                          pPessoa + '&' +
                                          'pCpfTitular=' + cpfReal + '&' +
                                          'pNomeTitular=' + codigo_usuarioNome + '&' +
                                          'pTipoInclusao=2&pRN=N&pUnidade=2&' +
                                          'pCodigoContrato=' + codigo_usuarioId[:11] + '&' +
                                          pIdSessao + '&' +
                                          'pCpf=' + cpfReal + '&' +
                                          'pNomeDep=' + nome)
                            ''' + '&' +
                                          'pDataNasc=' + dtNascimento + '&' +
                                          'pEstadoCivil=' + estadoCivil + '&' +
                                          'pTipoDep=' + dependencia + '&' +
                                          'pDtCasamento=' + dataCasamento + '&' +
                                          'pSexo(M, F)=' + sexo + '&' +
                                          'pMae=' + nomeMae)
                                          '''
                            aba_ativa[f"K{celula.row}"] = "Ativo"
                            planilha.save(caminhoBD)
                            linhasExecutadas += 1
                            time.sleep(5)
                            print('Status: Finalizada \n')
                            time.sleep(5)
                    else:
                        aba_ativa[f"K{celula.row}"] = "Falta codigoAtivo"
                        planilha.save(caminhoBD)
                        linhasExecutadas += 1
                        time.sleep(5)
                        print(vida + ' || Aguardando gerar codigoAtivo\n')
                else:
                    print('break')
                    break
            codigo_usuarioId = 0
            codigo_usuarioNome = 0
        comecarComTitular += 1
        if comecarComTitular == 2:
            print('Titulares verificados, iniciando verificação de dependentes\n')
except NoSuchWindowException:
    print('A janela do Chorme foi fechada\n')
finally:
    navegador.quit()
    print("___Execução Finalizada___")
    print('Linhas executadas: ', linhasExecutadas)
    print('Linhas com erro: ', linhasErro)